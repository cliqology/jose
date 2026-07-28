from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from jose.models import Source, SourceImportRun, User

SECTION_CATEGORY = {
    "VC FIRM": "vc_portfolio",
    "JOBS NEWSLETTERS": "newsletter",
    "VC FELLOWSHIPS / TALENT NETWORKS": "talent_network",
    "THE A16Z SPEEDRUN TALENT NETWORK": "talent_network",
}

HEADER_ROWS = {
    "JOB AGGREGATOR URL",
    "URL",
}


@dataclass
class ImportRowOutcome:
    row_number: int
    name: str
    url: str
    category: str | None
    action: str
    reason: str | None = None


def classify_workbook(
    session: Session, user: User, source: str | Path | BinaryIO
) -> list[ImportRowOutcome]:
    workbook = load_workbook(filename=source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    current_category: str | None = None
    seen_urls: set[str] = set()
    outcomes: list[ImportRowOutcome] = []

    for row_number, (name_cell, url_cell) in enumerate(
        sheet.iter_rows(min_col=1, max_col=2, values_only=True), start=1
    ):
        name = str(name_cell).strip() if name_cell else ""
        url = str(url_cell).strip() if url_cell else ""
        upper_name = name.upper()

        if upper_name in SECTION_CATEGORY:
            current_category = SECTION_CATEGORY[upper_name]
            continue

        if not url or url.upper() in HEADER_ROWS or not url.startswith(("http://", "https://")):
            outcomes.append(
                ImportRowOutcome(row_number, name, url, current_category, "skip", "No usable URL")
            )
            continue

        if url in seen_urls:
            outcomes.append(
                ImportRowOutcome(
                    row_number,
                    name,
                    url,
                    current_category,
                    "flag",
                    "Duplicate URL already seen earlier in this workbook",
                )
            )
            continue
        seen_urls.add(url)

        if current_category is None:
            outcomes.append(
                ImportRowOutcome(
                    row_number,
                    name,
                    url,
                    None,
                    "flag",
                    "URL found before any recognized section header",
                )
            )
            continue

        existing = session.scalar(
            select(Source).where(Source.user_id == user.id, Source.url == url)
        )
        action = "update" if existing else "create"
        outcomes.append(ImportRowOutcome(row_number, name, url, current_category, action))

    return outcomes


def apply_outcomes(session: Session, user: User, outcomes: list[ImportRowOutcome]) -> None:
    for outcome in outcomes:
        if outcome.action == "create":
            enabled = outcome.category == "vc_portfolio"
            session.add(
                Source(
                    user_id=user.id,
                    name=outcome.name or outcome.url,
                    url=outcome.url,
                    category=outcome.category,
                    portfolio_firm=outcome.name if outcome.category == "vc_portfolio" else None,
                    adapter="auto",
                    enabled=enabled,
                    notes="Imported from spreadsheet",
                )
            )
        elif outcome.action == "update":
            existing = session.scalar(
                select(Source).where(Source.user_id == user.id, Source.url == outcome.url)
            )
            assert existing is not None
            existing.name = outcome.name or existing.name
            existing.category = outcome.category or existing.category
            # `enabled` is intentionally left untouched on update: once a source
            # exists, its enabled state is user-controlled, not spreadsheet-controlled.


def summarize(outcomes: list[ImportRowOutcome]) -> dict[str, int]:
    counts = {"create": 0, "update": 0, "skip": 0, "flag": 0}
    for outcome in outcomes:
        counts[outcome.action] += 1
    return counts


def commit_import(
    session: Session, user: User, source: str | Path | BinaryIO, filename: str
) -> SourceImportRun:
    outcomes = classify_workbook(session, user, source)
    apply_outcomes(session, user, outcomes)
    counts = summarize(outcomes)
    flagged_rows = [
        {
            "row_number": outcome.row_number,
            "name": outcome.name,
            "url": outcome.url,
            "reason": outcome.reason,
        }
        for outcome in outcomes
        if outcome.action == "flag"
    ]
    run = SourceImportRun(
        user_id=user.id,
        filename=filename,
        created_count=counts["create"],
        updated_count=counts["update"],
        skipped_count=counts["skip"],
        flagged_count=counts["flag"],
        flagged_rows=flagged_rows,
    )
    session.add(run)
    session.commit()
    session.refresh(run)
    return run
